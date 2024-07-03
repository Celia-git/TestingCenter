import datetime as dt
import openpyxl
import openpyxl.utils
import openpyxl.utils.exceptions

Visitors = {
	"Student":{"A Number":"","Name":"", "Courses":[], "Visit Dates":[]},
	"Date":{"Date":"", "A Number":"", "Testing":True, "Course":"", "Section":"", "Calc #":0, "Time In":"", "Time Out":""}
}

class Date:


    # Param path is location of xl Workbook
    def __init__(self, path):
        self.path = path

    # Return Dict visit data (key:Datetime object, val:Arr data)
    # Param range: date range to return [1, 7, 30]
    def load_data(self, date_start, date_range):
        # Open Workbook and sheet
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb["Date"]
        data = {}
        
        # Iterate Rows of Data while in range
        [start_row, end_row] = self.get_index(date_start, date_range, ws)
        
        if not (end_row<=start_row):
            return data
        
        r = start_row
        while r >= end_row:
            date = ws.cell(row= r, column= 1).value
            # Iterate columns, add data to values
            vals = []
            col = 2
            entry = ws.cell(row=r, column=col).value
            while col <= len(Visitors["Date"]):
                # if entry is time, keep only hours and minutes
                entry = str(entry)
                if ":" in entry:
                    time=entry.split(":")
                    entry = "%s:%s"%(time[0], time[1])
                elif "=" in entry:
                    entry = entry.strip("()").strip("=")
                elif "\n" in entry:
                    entry = entry.strip()
                vals.append(entry)
                col +=1
                entry = ws.cell(row=r, column=col).value
            
            if date in data.keys():
                # add row data to dict at key
                data[date].append(vals)
            else:
                # Create new entry
                data[date] = [vals]
            r-=1    

        wb.close()
        return data
  
    # Return start and end index of the rows in range
    def get_index(self, start_date, drange, ws):
        start_date_idx = self.get_int_date(start_date)
        r = 2
        
        # Get index of starting date ROW: the index of the lowest date on the sheet
        while True:
            date_idx = self.get_int_date(ws.cell(row=r, column=1).value)
            # If date_idx > start_date_idx
            if (self.compare(date_idx, start_date_idx)==1) or (date_idx == None):
                break
            else:
                r += 1
        r -= 1
        start_row = r
        
        # Get tuple last Date in range
        end_date_idx = ()
        new_date = start_date_idx[0] - drange
        if new_date > 0:
            end_date_idx = (new_date, start_date_idx[1])
        else:
            end_date_idx = (new_date+365, start_date_idx[1]-1)
            
        # Get Index of Last row in date range: index of the lowest date on the sheet 
        while r >= 2:
            date_idx = self.get_int_date(ws.cell(row=r, column=1).value)
            comp = self.compare(date_idx, end_date_idx)
            if (comp==-1) or (date_idx==None):
                r += 1
                break
            elif (comp==0):
                # Iterate all rows with matching dates
                while self.compare(date_idx, end_date_idx)==0:
                    r -= 1
                    date_idx = self.get_int_date(ws.cell(row=r, column=1).value)
            else:
                r -=1
        if r < 2: r = 2
        return [start_row, r]
             
            
        
    # Return an int which represent's the date's position in year
    def get_int_date(self, string_date):
        if (string_date is None) or (string_date == "Date"):
            return None

        # If date was changed to datetime obj in xl file
        if type(string_date) == dt:
            string_date = "%s/%s/%s"%(string_date.month, string_date.day, string_date.year)
           
        dates = string_date.split("/")
        month = self.month(int(dates[0]))*(int(dates[0])-1)
        day = int(dates[1])
        # Get int x: years after 2020
        year = int(dates[2][-2:])

        return (month+day, year)
  
    # Save new data to file
    # Param Data: dictionary of values
    def save_data(self, data):
        # Open Workbook and sheet
        wb = openpyxl.load_workbook(self.path, read_only=False)
        ws = wb["Date"]

        max_row = self.get_max_row(ws)
        col = 1
        
        for key in data.keys():
            category = ws.cell(row=1, column=col).value
            if key==category:
                ws.cell(row=max_row, column=col).value = data[key]
            col +=1

        try:
            wb.save(self.path)
            wb.close()
            return 0
        except IOError:
            wb.close()
            return 1

    # Save checkout time in visit, Return 1 for error
    def save_student_checkout(self, date, a_num, time_in, time_out):
        # Open Workbook and sheet

        try:
            wb = openpyxl.load_workbook(self.path, read_only=False)
            ws = wb["Date"]
        except (IOError, openpyxl.utils.exceptions.InvalidFileException) as exception:
            return (1, exception.__str__())

        max_row = self.get_max_row(ws)
        row = self.get_max_row(ws)
        # Find row with matching date
        while row > 0:
            date_value = ws.cell(row=row, column=1).value
            if date_value == date:
                # Find row with matching Anumber and in-time
                a_num_value = ws.cell(row=row, column=2).value
                if a_num_value == a_num:
                    time_in_value = ws.cell(row=row, column=7).value
                    if time_in_value == time_in:
                        ws.cell(row=row, column=8).value = time_out

            row -= 1
        
        try:
            wb.save(self.path)
            wb.close()
            return (0, 0)
        except (IOError, openpyxl.utils.exceptions.InvalidFileException) as exception:
            wb.close()
            return (1, exception.__str__())


    # Return All Visits for A number at date
    # visits = {"date":[[visit1],[visit1]], [[visit2],[visit2]]}
    def get_visits(self, a_num):
        # Open Workbook and sheet
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True)
            ws = wb["Date"]
        except (IOError, openpyxl.utils.exceptions.InvalidFileException) as exception:
            return (1, exception.__str__())

        r=2
        visits = {"Date": [["Testing", "Course", "Section", "Calc #", "Time In", "Time Out"]]}
        max_row = self.get_max_row(ws)
        while r <= max_row:
            entry_date = ws.cell(row=r, column=1).value
            entry_a = ws.cell(row=r, column=2).value
            vals = []
            if (entry_a==a_num):
                col = 3
                entry = ws.cell(row=r, column=col).value
                while col <= len(Visitors["Date"]):
                    # if entry is time, keep only hours and minutes
                    entry = str(entry)
                    if ":" in entry:
                        time = entry.split(":")
                        entry = "%s:%s" % (time[0], time[1])
                    vals.append(entry)
                    col += 1
                    entry = ws.cell(row=r, column=col).value
            # Add this data to student's list of visits
            if entry_date in visits:
                visits[entry_date].append(vals)
            # Create a new list of visits at entry date
            elif len(vals)>0:
                visits[entry_date] = [vals]
            r+=1
        
        try:
            
            wb.close()
            return visits
        except (IOError, openpyxl.utils.exceptions.InvalidFileException) as exception:
            wb.close()
            return (1, exception.__str__())

    def get_max_row(self, ws):
        for max_row, row in enumerate(ws, 1):
            if all(c.value is None for c in row):
                break
        return max_row

    # Compare two 'dates': (date_idx, year)
    # Return 1 if one is greater, -1 if less, 0 if equal
    def compare(self, date_one, date_two):
        if (date_two is None)or(date_one is None): return
        # Compare Years
        if date_one[1] > date_two[1]: return 1
        elif date_one[1] < date_two[1]: return -1
        else:
            if date_one[0] > date_two[0]: return 1
            elif date_one[0] < date_two[0]: return -1
        return 0

    # Return amount of days in the month
    def month(self, m):
        switcher = {
            1:31,
            2:28,
            3:31,
            4:30,
            5:31,
            6:30,
            7:31,
            8:31,
            9:30,
            10:31,
            11:30,
            12:31
        }
        return switcher.get(m, 0)