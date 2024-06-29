import openpyxl

class Student:

    def __init__(self, path, aNum):
        self.path = path
        self.aNum = aNum
        self.name = ""
        self.courses = []
        self.visits = []
        self.load_data(aNum)

        
    # Load student data from file, set instance vars
    def load_data(self, aNum):
        # Open Workbook and sheet
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb["Student"]

        # Iterate rows until student's A nuber is reached
        r = 2
        col = 1
        entry = ws.cell(row=r, column=col).value

        while entry is not None:
            if entry==self.aNum:
                # iterate columns and save their values
                while True:
                    
                    value = ws.cell(row=r, column=col).value
                    category = ws.cell(row=1, column=col).value

                    if category is None: break

                    if value is None:
                        col += 1
                        continue

                    elif category=="A Number":
                        self.set_aNum(value)
                        
                    elif category=="Name":
                        self.set_name(value)
                        
                    elif category=="Courses":
                        # Break Value into smaller variables
                        return_values = []
                        course_array = value.split(",")
                        for c in range(len(course_array)):
                            cour=course_array[c].split(":")[0].strip()
                            sec=course_array[c].split(":")[1].strip()
                            return_values.append([cour, sec.strip()])
                        self.set_courses(return_values)
                        
                    elif category=="Visit Dates":
                        all_dates = value.split(",")
                        for date in all_dates:
                            self.add_visit(date.strip())
                    
                    col += 1
                    
                
                break
                
            r += 1
            entry = ws.cell(row=r, column=col).value

        wb.close()
 
    # Return student values
    # Return String aNum
    def get_aNum(self):
        return self.aNum

    # Return String name
    def get_name(self):
        return self.name

    # 2D Array of courses and sections
    def get_courses(self):
        return self.courses
        
    # Return array of visits
    def get_visits(self):
        return self.visits
        
    # Set Student Values
    # param: String
    def set_aNum(self, aNum):
        self.aNum = aNum
        
    # param: String
    def set_name(self, name):
        self.name = name
    
    # param: 2D array
    def set_courses(self, courses):
        self.courses = courses
        
    # param: String Date
    def add_visit(self, visit):
        self.visits.append(visit)
    

    # Save Instance Vars to file
    def save_data(self):
        # Open Workbook and sheet
        wb = openpyxl.load_workbook(self.path, read_only=False)
        ws = wb["Student"]

        # Iterate rows until student's A nuber is reached
        r = 1
        entry = ws.cell(row=r, column=2).value
        a_exists = False
        while entry is not None:
            if entry==self.aNum:
                self.set_row_values(r, ws)
                a_exists=True
                break
            r += 1
            entry = ws.cell(row=r, column=1).value
        
        # If no entry exists at A Number, create a new one
        if not a_exists:
            self.set_row_values(r, ws)

        try:
            wb.save(self.path)
            wb.close()
            return 0
        except IOError:
            wb.close()
            return 1

            
    def set_row_values(self, row, ws):
        # iterate columns and set their values
        col = 1
        category = ws.cell(row=1, column=col).value
        while category is not None:

            if category=="A Number":
                ws.cell(row=row, column=col).value = self.aNum
                
            elif category=="Name":
                ws.cell(row=row, column=col).value = self.name
                
            elif category=="Courses":
                save_string = ""
                # Convert 2D array to string
                i=0
                for course in self.courses:
                    punct=","
                    if i == len(self.courses)-1: punct=""
                    save_string += "%s:%s%s"%(course[0], course[1], punct)
                    i+=1
                ws.cell(row=row, column=col).value = save_string
                
            elif category=="Visit Dates":
                save_string = ""
                i=0
                for visit in self.visits:
                    punct=","
                    if i == len(self.courses)-1: punct=""
                    save_string += "%s%s"%(visit, punct)
                ws.cell(row=row, column=col).value=save_string

            col += 1
            category = ws.cell(row=1, column=col).value


    # Param array [Course,Section]
    # Return boolean whether student has course
    def has_course(self, course):
        return (course in self.courses)