
import tkinter as tk
from tkinter import ttk
import ctypes
from datetime import datetime
import os
import copy
import openpyxl
from Date import Date, Visitors
from Student import Student

data_path = "data\\Visitors.xlsx"
courses_path = "data\\Courses.txt"

FONT = ("calibre", 12, "bold")
backgrounds = {"blue":"#A8DADC", "grey":"#CFD8DC", "white":"#ECEFF1"}

class App(tk.Tk):
    def __init__(self):
        # If Data files do not exist, create them
        if not os.path.exists(data_path):
            wb = openpyxl.Workbook()
            # Create Each Sheet
            for key in Visitors.keys():
                sheet = wb.create_sheet(key)
                sheet.title = key
                # Add Each column
                col = 1
                for val in Visitors[key].keys():
                    sheet.cell(row=1, column=col).value = val
                    col +=1
            wb.save(filename = data_path)
            

        # create tkinter window
        super().__init__()
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
        self.title('Student Scanner')
        self.geometry("900x600")

        # popup the Window
        container_frame = ContainerFrame(self)
        container_frame.pack(padx=50, pady=50, fill=tk.BOTH, side=tk.TOP) 
        container_frame.swap(0)


# Control Frame: Holds all sub frames
class ContainerFrame(tk.Frame):
    def __init__(self, container):
        super().__init__(container)

        # Get Current Date
        self.now = datetime.now()
        self.date = self.now.strftime('%m/%d/%y')

        # Create time and date widgets
        self.time_label = tk.Label(self, font=FONT)
        self.date_label = tk.Label(self, font=FONT, text=self.date)
        self.back_but = tk.Button(self, font=FONT, text="<--", command=lambda : self.swap(0))
        self.time_label.pack(side=tk.TOP)
        self.date_label.pack(side=tk.TOP)
        self.back_but.pack(side=tk.BOTTOM)
        # Create frames
        
        self.a_num = tk.StringVar()
        self.default_frame = DefaultFrame(self)
        self.student_frame = StudentFrame(self)
        self.date_frame = DateFrame(self)
        self.time()

    # Refresh time clock
    def time(self):
        self.now = datetime.now()
        string = self.now.strftime('%H:%M')
        self.time_label.config(text=string)
        self.time_label.after(60000, self.time)

    # Default Frame
    def load_default_frame(self):
        self.back_but["state"] = "disabled"
        self.student_frame.pack_forget()
        self.student_frame.discard(True)
        self.date_frame.clear_display()
        self.date_frame.pack_forget()
        self.a_num.set("")
        self.default_frame.entry.focus_set()
        return self.default_frame

    # Student Frame
    def load_student_frame(self):
        self.back_but["state"] = "normal"
        self.student_frame.load_student(self.a_num.get())
        self.date_frame.pack_forget()
        self.default_frame.pack_forget()
        return self.student_frame

    # Date frame
    def load_date_frame(self):
        self.back_but["state"] = "normal"
        self.default_frame.pack_forget()
        self.student_frame.pack_forget()
        return self.date_frame

    # Show correct frame
    def swap(self, frame_idx):
        switch = {
            0: self.load_default_frame,
            1: self.load_student_frame,
            2: self.load_date_frame

        }
        next = switch.get(frame_idx, "Invalid Window Index")
        top_frame = next()
        top_frame.pack(padx=25, pady=25, side=tk.TOP, fill=tk.BOTH, expand=True)
        
    # Returns true if string is A00 followed by six digits
    def is_valid_Anumber(self):
        string_anumber = self.a_num.get()
        if string_anumber.startswith("A00") and len(string_anumber)==9:
            i = 3
            while i<9:
                if not string_anumber[i].isdigit():
                    return False
                i+=1
            return True
        return False


# Default Program window
class DefaultFrame(tk.Frame):
    def __init__(self, container):
        super().__init__(container)
        
        # Configure Frame Settings
        self.container = container
        self.columnconfigure((0,1,2,4,5), weight=1)
        self.columnconfigure(3, weight=2)
        self.rowconfigure((0,1,2), weight=1)

        # Create widgets
        self.error_label = tk.Label(self, font=FONT, text="")
        a_label = tk.Label(self, text="A Number:", font=FONT, pady=10)
        v_label = tk.Label(self, text="View Logs by Date:", font=FONT, pady=10)
        self.entry = tk.Entry(self, textvariable=self.container.a_num, font=FONT)
        view_but = tk.Button(self, text="Logs", font=FONT, command=self.view, pady=10)
        submit_but = tk.Button(self, text="Go", font=FONT, command=self.submit, pady=10)
        self.entry.bind("<Return>", self.submit)

        # Place widgets
        a_label.grid(row=0, column=2)
        v_label.grid(row=2, column=2)
        self.entry.grid(row=0, column=3)
        view_but.grid(row=2, column=3)
        submit_but.grid(row=0, column=4)
        self.error_label.grid(row=3, column=1, columnspan=3)

        

    ####### Commands #######

    # Open Student Log
    def submit(self, *event):
        if self.container.is_valid_Anumber():
            self.error_label.config(text="")
            self.container.swap(1)
        else:
            self.error_label.config(text="Error: %s: not a valid A Number\nEnter a student ID which starts with A00 and ends with six digits" % self.container.a_num.get())

    # View date logs
    def view(self):
        self.error_label.config(text="")
        self.container.swap(2)


# Show student log window
class StudentFrame(tk.Frame):
    def __init__(self, container):
        super().__init__(container)
        
        #Variables
        self.container = container
        self.name = tk.StringVar()
        self.course=tk.StringVar()
        self.section=tk.StringVar()
        self.calc=tk.IntVar()
        self.test=tk.BooleanVar()
        
        # Load Courses from txt
        file = open(courses_path)
        self.course_list = file.readlines()
        file.close()
        self.all_courses = []
        self.all_secs = {}
        for line in self.course_list:
            c = line.split(":")[0]
            self.all_courses.append(c)
            self.all_secs[c] = []
            for x in line.split(":")[1].split(",") : self.all_secs[c].append(x.strip("\n"))
        
        # Create Frames and Scrolling
        self.log_visit = tk.Frame(self, padx=20, pady=10)
        self.log_visit.pack(side="top", fill=tk.BOTH)
        self.cmd_frame = tk.Frame(self, padx=20, pady=20)
        self.cmd_frame.pack(side="top", fill=tk.X)
        self.display_canvas = tk.Canvas(self)
        self.display_frame = tk.Frame(self.display_canvas)
        self.sb = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.display_canvas.yview)
        self.display_canvas.configure(yscrollcommand=self.sb.set)

        self.sb.pack(side="right", fill="y")
        self.display_canvas.pack(side="top", fill=tk.BOTH, expand=True)
        self.display_canvas.create_window((4,4), window=self.display_frame, anchor="center", tags="self.display_frame")
        self.display_frame.bind("<Configure>", self.onFrameConfigure)

        # Create log_visit widgets
        self.log_visit.columnconfigure((0,1,2), weight=1)
        self.log_visit.rowconfigure((0,1,2, 3), weight=1)
        a_label = tk.Label(self.log_visit, font=FONT, text="A Number: ", padx=5, pady=5)
        a_value = tk.Label(self.log_visit, font=FONT, textvariable=self.container.a_num, padx=5, pady=5)
        n_label = tk.Label(self.log_visit, font=FONT, text="Name: ", padx=5, pady=5)
        n_value = tk.Entry(self.log_visit, font=FONT, textvariable=self.name)
        cor_label = tk.Label(self.log_visit, font=FONT, text="Student's Courses: ", padx=5, pady=5)
        self.cor_value = ttk.Combobox(self.log_visit, font=FONT, values=self.all_courses, textvariable=self.course, width=10)
        sec_label = tk.Label(self.log_visit, font=FONT, text="Section: ", padx=5, pady=5)
        self.sec_value = ttk.Combobox(self.log_visit, font=FONT, values=[], textvariable=self.section, width=5)
        cal_label = tk.Label(self.log_visit, font=FONT, text="Calculator #", padx=5, pady=5)
        cal_value = tk.Spinbox(self.log_visit, font=FONT, from_ = 1, to = 90,
                                     increment = 1, textvariable=self.calc, width=3)
        test_value = tk.Checkbutton(self.log_visit, font=FONT, text="Testing", variable=self.test)
        
        a_label.grid(row=0, column=0, columnspan=2)
        a_value.grid(row=0, column=2, columnspan=2)
        n_label.grid(row=1, column=0, columnspan=2)
        n_value.grid(row=1, column=2, columnspan=2)
        cor_label.grid(row=2, column=0)
        self.cor_value.grid(row=2, column=1)
        self.cor_value.bind("<<ComboboxSelected>>", self.set_section_values)
        sec_label.grid(row=2, column=2)
        self.sec_value.grid(row=2, column=3)
        cal_label.grid(row=3, column=0)
        cal_value.grid(row=3, column=1)
        test_value.grid(row=3, column=2, columnspan=2)

        # Create Command Frame widget
        self.cmd_frame.columnconfigure((0,1,2), weight=1)
        self.cmd_frame.rowconfigure((0,1,2), weight=1)
        discard_button = tk.Button(self.cmd_frame, font=FONT, text="Discard Changes", command=self.discard, padx=10)
        self.error_label = tk.Label(self.cmd_frame, font=FONT, text="")
        save_button = tk.Button(self.cmd_frame, font=FONT, text="Save Changes", command=self.save, padx=10)
        visit_label = tk.Label(self.cmd_frame, font=FONT, text="Past Visits:", padx=10, pady=10)
        discard_button.grid(row=0, column=0)
        self.error_label.grid(row=0, column=1)
        save_button.grid(row=0, column=2)
        visit_label.grid(row=2, column=1, sticky="EW")

    # Load Student data from file
    def load_student(self, a_num):
        self.student = Student(data_path, a_num)
        self.date_log = Date(data_path)
        self.container.a_num.set(a_num)
        self.name.set(self.student.get_name())
        self.test.set(True)
        
        # Write Student Courses and Sections to Entry Fields
        student_courses = self.student.get_courses()
        if student_courses:
            new_cor_vals = []
            for c in student_courses:
                new_cor_vals.append(c[0])
            new_cor_vals.append("---------------------")
            for c in self.all_courses:
                new_cor_vals.append(c)

            self.cor_value.config(values=new_cor_vals)
            self.cor_value.set(new_cor_vals[0])
            self.set_section_values("", student_courses[0][1])

        # Write past visits to Display_frame
        visits = self.date_log.get_visits(a_num)
        if type(visits)==dict:
            self.display_frame.columnconfigure((0, len(visits.keys())), weight=1)
            row=1
            column=0
            for date in visits.keys():
                color = "white"
                if isinstance(date, str):
                    date_label = tk.Label(self.display_frame, font=FONT, text=date, bg=backgrounds["blue"])
                else:
                    date_label = tk.Label(self.display_frame, font=FONT, text="%s/%s/%s"%(date.month, date.day, date.year), bg=backgrounds["blue"])
                date_label.grid(row=row, column=0, columnspan=len(visits[date][0]), sticky="EW", ipady=10)
                # Iterate Visits at Date
                for v in range(len(visits[date])):
                    row +=1
                    if v%2!=0:color="grey"
                    # Iterate entries in Visit
                    for entry in visits[date][v]:
                        if "=" in entry:
                            entry = entry.strip("()").strip("=")
                        this_entry = tk.Label(self.display_frame, font=FONT, text=entry, width=len(entry)+4, justify=tk.CENTER, bg=backgrounds[color])
                        this_entry.grid(row=row, column=column, sticky="EW")
                        column +=1
                    column = 0
                row += 1
       
    def set_section_values(self, _event, *default):
        self.sec_value.delete(0, 'end')
        if self.course.get() in self.all_courses:
            self.sec_value.config(values=self.all_secs[self.course.get()])
            self.sec_value.set(default)
        else:
            self.sec_value.config(values=[])
    
    # Reset the scroll region
    def onFrameConfigure(self, event):
        self.display_canvas.configure(scrollregion=self.display_canvas.bbox("all"))

    # Save Entries, go to default menu
    def save(self):

        # Gather All Values for Date Log
        date_dict = copy.deepcopy(Visitors["Date"])
        for key in date_dict.keys():
            if key=="Date":
                date_dict[key] = self.container.date
            elif key=="A Number":
                date_dict[key] = self.container.a_num.get()
            elif key=="Testing":
                date_dict[key] = self.test.get()
            elif key=="Course":
                date_dict[key] = self.course.get()
            elif key=="Section":
                date_dict[key] = self.section.get()
            elif key=="Calc #":
                date_dict[key] = self.calc.get()
            elif key=="Time In":
                date_dict[key] = self.container.now.strftime('%H:%M')
        if self.date_log is None:
            self.date_log = Date(data_path)

        # Set All Student values
        self.student.set_aNum(self.container.a_num.get())
        self.student.set_name(self.name.get())
        self.student.add_visit(self.container.date)
        # Validate courses input
        invalid = False
        if not (self.course.get() in self.all_courses):
            invalid=True
        elif not (self.section.get() in self.all_secs[self.course.get()]):
            invalid=True
        if invalid:
            self.error_label.config(text="Error: %s: %s not found\nEnter a different course or add it to %s" % (self.course.get(), self.section.get(),self.courses_path))
            return
        # If Student Doesn't Have course, add it to their list
        if not self.student.has_course([self.course.get(),self.section.get()]):
            new_courses = [[self.course.get(), self.section.get()]]
            old_courses = self.student.get_courses()
            for c in old_courses:
                new_courses.append(c)
            self.student.set_courses(new_courses)

        # Save all Values
        return_val = self.date_log.save_data(date_dict) + self.student.save_data()
        if return_val==0:
            self.discard()
        else:
            # Show error Message
            self.error_label.config(text="IOError when writing to file. \nMake sure %s is closed"%data_path)
    
    # Reset Entry Fields, go to default menu
    def discard(self, *stop):
        self.container.a_num.set("")
        self.name.set("")
        self.course.set("")
        self.section.set("")
        self.calc.set(0)
        self.test.set(True)
        self.error_label.config(text="")
        # Reset display frame
        for label in self.display_frame.grid_slaves():
            label.grid_forget()
        if not stop:
            self.container.swap(0)

# Show date log window
class DateFrame(tk.Frame):
    def __init__(self, container):
        super().__init__(container)
        
        # Variables
        self.container = container
        self.now = container.now
        self.search_date = tk.StringVar()
        self.search_date.set(self.now.strftime("%m/%d/%y"))
        self.days_back = tk.IntVar()
        self.days_back.set(7)


        # Create Frames + Canvas + Scrollbar
        self.entry = tk.Frame(self, padx=25, pady=25)
        self.entry.pack(side="top", fill=tk.X, expand=True, padx=40, pady=15)

        self.display_canvas = tk.Canvas(self, **{"highlightthickness":0})
        self.display_canvas.pack(side="left", fill="both", expand=True)
        self.display_frame = tk.Frame(self.display_canvas)
        self.sb = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.display_canvas.yview)
        self.sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.display_canvas.configure(yscrollcommand=self.sb.set)
        self.display_frame.bind("<Configure>", self.onFrameConfigure)
        self.display_canvas.create_window((4,4), window=self.display_frame, anchor=tk.NW, tags="self.display_frame")
        self.error_label = tk.Label(self.display_frame, font=FONT, text="")
        self.error_label.grid(row=3, column=0)

        # Create Entry Widgets
        labelA = tk.Label(self.entry, font=FONT, text="View visits ")
        days_back_spin = tk.Spinbox(self.entry, font=FONT, from_ = 1, to = 90,
                                     increment = 1, textvariable = self.days_back, width=3)
        labelB = tk.Label(self.entry, font=FONT, text="days from ")
        self.search_date_entry = tk.Entry(self.entry, font=FONT, textvariable=self.search_date, width=8)
        self.search_date_entry.bind("<Return>", self.display)
        go_button = tk.Button(self.entry, text="GO", font=FONT, command= self.display)
        
        labelA.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        days_back_spin.pack(side=tk.LEFT, padx=5)
        labelB.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        self.search_date_entry.pack(side=tk.LEFT, padx=5)
        go_button.pack(side=tk.LEFT, padx=5, fill=tk.X)
        
        
    # Reset the scroll region
    def onFrameConfigure(self, event):
        self.display_canvas.configure(scrollregion=self.display_canvas.bbox("all"))

    # Generate Display labels
    def display(self, *event):

        # Clear Display Values, Validate Entries
        self.clear_display()
        entered_date=self.search_date.get()
        entered_days=self.days_back.get()
        if not self.is_valid_date(entered_date):
            entered_date = self.now.strftime("%m/%d/%y")
        
        # Get All entries from 'days_back' days from 'search_from'
        self.date_log = Date(data_path)
        loaded = self.date_log.load_data(entered_date, entered_days)
        
        # Row 0: Columns
        column = 0
        for val in Visitors["Date"].keys():
            if val=="Date":continue
            lab = tk.Label(self.display_frame, font=FONT, text=val, padx=10,pady=25)
            lab.grid(row=0, column=column)
            self.display_frame.columnconfigure(column, weight=1)
            column+=1
        
        # Iterate Dates
        row = 1
        column = 0
        for date in loaded.keys():
            if date is None:
                continue
            if isinstance(date, str):
                date_label = tk.Label(self.display_frame, font=FONT, text=date, bg=backgrounds["blue"])
            else:
                date_label = tk.Label(self.display_frame, font=FONT, text="%s/%s/%s"%(date.month, date.day, date.year), bg=backgrounds["blue"])
            date_label.grid(row=row, column=0, columnspan=len(loaded[date][0]), sticky="EW", ipady=10)
            
            row += 1
            # Iterate Visits at Date
            for v in range(len(loaded[date])):
                color = ["grey", "white"][int (v%2 == 0)]
                # Iterate entries in Visit
                for entry in loaded[date][v]:
                    # Add button to clock the checkout
                    if column == 6 and entry == "None":
                        check_out = tk.Button(self.display_frame, font=FONT, text="Checkout", command=lambda this_date=date, this_anum=loaded[date][v][0], time_in=loaded[date][v][5]: self.checkout_student(this_date, this_anum, time_in))
                        check_out.grid(row=row, column=column, sticky="EW")
                    # Add button to the A Number entry
                    elif column == 0:
                        anum_but = tk.Button(self.display_frame, font=FONT, text=entry, command=lambda anum=entry: self.show_student(anum))
                        anum_but.grid(row=row, column=column, sticky="EW")
                    # Add label to normal entries
                    else:
                        this_entry = tk.Label(self.display_frame, font=FONT, text=entry, width=len(entry)+4, justify=tk.CENTER, bg=backgrounds[color])
                        this_entry.grid(row=row, column=column, sticky="EW")
                    column +=1
                column = 0
                
                row += 1
            row += 1
            
    def show_student(self, anum):
        self.container.a_num = tk.StringVar(self, anum)
        self.container.swap(1)

    # End visit by logging end
    def checkout_student(self, date, a_num, time_in):
        # Save all Values
        return_val = self.date_log.save_student_checkout(date, a_num, time_in, self.container.now.strftime('%H:%M'))     
        if return_val[0]==0:
            # Refresh page
            self.display()
        else:
            # Show error Message
            self.error_label.config(text="%s. \nMake sure %s is closed and a valid filepath"% (return_val[1], data_path))

    def clear_display(self):
        for label in self.display_frame.grid_slaves():
            label.grid_forget()
                
    # Returns true if date is formatted correctly
    def is_valid_date(self, string_date):
        date_format = '%m/%d/%y'
        try:
            datetime.strptime(string_date, date_format)
            return True
        except ValueError:
            return False                    

        

